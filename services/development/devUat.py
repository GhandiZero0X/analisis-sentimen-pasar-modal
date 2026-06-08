"""
Analisis Hasil Survey Pengguna (UAT) - StockSenseID
====================================================
Menghasilkan visualisasi lengkap untuk bagian 4.10.2 Survey Pengguna:
  1. Grafik rata-rata skor per pertanyaan (bar chart)
  2. Grafik radar rata-rata per kategori
  3. Grafik distribusi skor keseluruhan (pie + stacked bar)
  4. Heatmap skor per responden
  5. Tabel ringkasan statistik deskriptif (Excel)

Struktur folder:
  services/
    development/
      devUat.py                          ← file ini
      dev_database/
        10_Uat/
          Kuesioner UAT ... (Jawaban) (1).xlsx   ← input
          fig1_bar_per_pertanyaan.png            ← output
          fig2_radar_per_kategori.png            ← output
          fig3_distribusi_skor.png               ← output
          fig4_heatmap.png                       ← output
          tabel_statistik_uat.xlsx               ← output

Penggunaan (jalankan dari folder services/development/):
  python devUat.py
"""

import os
import warnings

import matplotlib
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
matplotlib.rcParams["font.family"]       = "DejaVu Sans"
matplotlib.rcParams["axes.spines.top"]   = False
matplotlib.rcParams["axes.spines.right"] = False


# ─── 0. Path ───────────────────────────────────────────────────────────────────

# Direktori tempat devUat.py berada  →  services/development/
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Folder input & output  →  services/development/dev_database/10_Uat/
UAT_DIR = os.path.join(BASE_DIR, "dev_database", "10_Uat")

FILE_PATH = os.path.join(
    UAT_DIR,
    "Kuesioner UAT Sistem Analisis Sentimen Saham (StockSenseID) (Jawaban) (1).xlsx",
)

# Output disimpan di folder yang sama (10_Uat)
OUT_DIR = UAT_DIR


# ─── 1. Konfigurasi pertanyaan & kategori ─────────────────────────────────────

LIKERT_COLS = [
    "A1", "A2", "A3", "A4",
    "B1", "B2", "B3", "B4",
    "C1", "C2", "C3", "C4",
    "D1", "D2", "D3",
]

PERTANYAAN = {
    "A1": "Informasi sentimen membantu\nmemahami persepsi publik",
    "A2": "Data sentimen relevan\ndengan kondisi pasar",
    "A3": "Hasil analisis dapat dijadikan\nreferensi tambahan",
    "A4": "Memudahkan melihat perbedaan\nsentimen antar saham",
    "B1": "Navigasi untuk menemukan\ninformasi mudah",
    "B2": "Antarmuka mudah dipahami\ntanpa panduan khusus",
    "B3": "Memilih saham & melihat\nsentimen cepat dan efisien",
    "B4": "Sistem merespons setiap\ninteraksi dengan cepat",
    "C1": "Visualisasi sentimen mudah\ndibaca dan dipahami",
    "C2": "Label kategori sentimen\nsudah cukup jelas",
    "C3": "Tren sentimen dari waktu\nke waktu membantu pemahaman",
    "C4": "Informasi yang disajikan\nsudah lengkap dan mencukupi",
    "D1": "Puas dengan tampilan\ndan antarmuka sistem",
    "D2": "Puas dengan informasi\nsentimen yang dihasilkan",
    "D3": "Bersedia merekomendasikan\nsistem kepada orang lain",
}

KATEGORI = {
    "A": {
        "cols": ["A1", "A2", "A3", "A4"],
        "label": "Manfaat &\nRelevansi",
        "color": "#378ADD",
    },
    "B": {
        "cols": ["B1", "B2", "B3", "B4"],
        "label": "Kemudahan\nPenggunaan",
        "color": "#3BB27A",
    },
    "C": {
        "cols": ["C1", "C2", "C3", "C4"],
        "label": "Visualisasi &\nKelengkapan",
        "color": "#E09B2D",
    },
    "D": {
        "cols": ["D1", "D2", "D3"],
        "label": "Kepuasan\nKeseluruhan",
        "color": "#D85A30",
    },
}

CAT_COLORS = {
    col: info["color"]
    for info in KATEGORI.values()
    for col in info["cols"]
}


# ─── 2. Baca & hitung statistik ────────────────────────────────────────────────

def load_data(path: str) -> pd.DataFrame:
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"File tidak ditemukan:\n  {path}\n"
            "Pastikan nama file dan lokasi sudah benar."
        )
    df = pd.read_excel(path)
    missing = [c for c in LIKERT_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Kolom berikut tidak ditemukan di file: {missing}")
    return df[LIKERT_COLS].copy()


df_score  = load_data(FILE_PATH)
n_resp    = len(df_score)
means     = df_score.mean().round(2)
cat_means = {
    k: round(df_score[v["cols"]].values.mean(), 2)
    for k, v in KATEGORI.items()
}
overall   = round(df_score.values.mean(), 2)

print(f"{'─'*50}")
print(f"  Analisis UAT – StockSenseID")
print(f"{'─'*50}")
print(f"  Jumlah responden  : {n_resp}")
print(f"  Rata-rata overall : {overall}")
print()
print("  Rata-rata per kategori:")
for k, v in cat_means.items():
    lbl = KATEGORI[k]["label"].replace("\n", " ")
    print(f"    {k} – {lbl:<28}: {v}")
print()
print("  Rata-rata per pertanyaan:")
for col in LIKERT_COLS:
    print(f"    {col}: {means[col]}")
print(f"{'─'*50}")


# ─── Helper ────────────────────────────────────────────────────────────────────

def save(fig: plt.Figure, name: str) -> str:
    path = os.path.join(OUT_DIR, name)
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  Disimpan: {path}")
    return path


# ─── 3. Gambar 1 – Bar chart rata-rata per pertanyaan ─────────────────────────

fig, ax = plt.subplots(figsize=(14, 6))

bar_colors = [CAT_COLORS[c] for c in LIKERT_COLS]
bars = ax.bar(
    LIKERT_COLS, means.values,
    color=bar_colors, edgecolor="white", linewidth=0.8,
    width=0.65, zorder=3,
)

ax.axhline(overall, color="#555", linewidth=1.2, linestyle="--", zorder=2)
ax.text(
    len(LIKERT_COLS) - 0.4, overall + 0.04,
    f"Rata-rata = {overall}",
    fontsize=8.5, color="#555", va="bottom", ha="right",
)

for bar, val in zip(bars, means.values):
    ax.text(
        bar.get_x() + bar.get_width() / 2, val + 0.04,
        f"{val:.2f}", ha="center", va="bottom",
        fontsize=8.5, fontweight="bold",
    )

ax.set_ylim(1, 5.6)
ax.set_ylabel("Rata-rata Skor (1–5)", fontsize=10)
ax.set_title(
    f"Rata-rata Skor per Pertanyaan UAT\n(n = {n_resp} responden)",
    fontsize=12, fontweight="bold", pad=12,
)
ax.set_xticks(range(len(LIKERT_COLS)))
ax.set_xticklabels(LIKERT_COLS, fontsize=9)
ax.yaxis.set_major_locator(plt.MultipleLocator(1))
ax.grid(axis="y", alpha=0.3, zorder=0)

legend_patches = [
    mpatches.Patch(
        color=v["color"],
        label=f"Kat. {k} – {v['label'].replace(chr(10), ' ')}",
    )
    for k, v in KATEGORI.items()
]
ax.legend(
    handles=legend_patches, loc="lower right",
    fontsize=8, framealpha=0.9, edgecolor="#ddd",
)

plt.tight_layout()
out1 = save(fig, "fig1_bar_per_pertanyaan.png")


# ─── 4. Gambar 2 – Radar chart per kategori ───────────────────────────────────

cat_labels  = [v["label"] for v in KATEGORI.values()]
cat_vals    = [cat_means[k] for k in KATEGORI]
cat_clrs    = [v["color"]  for v in KATEGORI.values()]

n_cat          = len(cat_labels)
angles         = np.linspace(0, 2 * np.pi, n_cat, endpoint=False).tolist()
angles_closed  = angles + [angles[0]]
vals_closed    = cat_vals + [cat_vals[0]]

fig, ax = plt.subplots(figsize=(6, 6), subplot_kw={"polar": True})

ax.plot(angles_closed, vals_closed, "o-", color="#378ADD", linewidth=2, markersize=6)
ax.fill(angles_closed, vals_closed, color="#378ADD", alpha=0.18)

for ang, val, col in zip(angles, cat_vals, cat_clrs):
    ax.plot(ang, val, "o", color=col, markersize=9, zorder=5)
    ax.text(
        ang, val + 0.15, f"{val:.2f}",
        ha="center", va="center",
        fontsize=9, fontweight="bold", color=col,
    )

ax.set_xticks(angles)
ax.set_xticklabels(cat_labels, fontsize=9.5, fontweight="bold")
ax.set_ylim(1, 5)
ax.set_yticks([1, 2, 3, 4, 5])
ax.set_yticklabels(["1", "2", "3", "4", "5"], fontsize=7.5, color="#888")
ax.set_title(
    f"Rata-rata Skor per Kategori UAT\n(n = {n_resp} responden)",
    fontsize=11, fontweight="bold", pad=18,
)
ax.grid(color="#ccc", alpha=0.5)
ax.text(
    0, -1.7, f"Overall: {overall}",
    ha="center", fontsize=10, fontweight="bold", color="#333",
    transform=ax.transData,
)

plt.tight_layout()
out2 = save(fig, "fig2_radar_per_kategori.png")


# ─── 5. Gambar 3 – Pie + stacked bar distribusi skor ─────────────────────────

all_scores = df_score.values.flatten()
total_ans  = len(all_scores)
score_cnt  = {s: int((all_scores == s).sum()) for s in [1, 2, 3, 4, 5]}

labels_pie = [f"Skor {s}" for s in [5, 4, 3, 2, 1]]
sizes_pie  = [score_cnt[s] for s in [5, 4, 3, 2, 1]]
colors_pie = ["#3BB27A", "#378ADD", "#E09B2D", "#D85A30", "#E24B4A"]
explode    = (0.04, 0.02, 0, 0, 0)

fig, (ax_pie, ax_bar) = plt.subplots(1, 2, figsize=(11, 5))

wedges, _, autotexts = ax_pie.pie(
    sizes_pie, labels=None, colors=colors_pie,
    autopct=lambda p: f"{p:.1f}%" if p > 1 else "",
    startangle=90, explode=explode,
    wedgeprops={"edgecolor": "white", "linewidth": 1.5},
    textprops={"fontsize": 9},
)
for at in autotexts:
    at.set_fontweight("bold")
ax_pie.legend(
    wedges, [f"{l} ({c})" for l, c in zip(labels_pie, sizes_pie)],
    loc="lower left", fontsize=8.5, framealpha=0.9,
)
ax_pie.set_title(
    f"Distribusi Skor Keseluruhan\n(semua pertanyaan, n = {total_ans} jawaban)",
    fontsize=10.5, fontweight="bold",
)

score_matrix     = {}
score_colors_bar = {5: "#3BB27A", 4: "#378ADD", 3: "#E09B2D", 2: "#D85A30", 1: "#E24B4A"}
cat_labels_short = [v["label"].replace("\n", " ") for v in KATEGORI.values()]

for k, v in KATEGORI.items():
    vals = df_score[v["cols"]].values.flatten()
    score_matrix[k] = {s: int((vals == s).sum()) for s in [1, 2, 3, 4, 5]}

bottom = np.zeros(len(KATEGORI))
for s in [1, 2, 3, 4, 5]:
    vals = np.array([score_matrix[k][s] for k in KATEGORI])
    bars_h = ax_bar.barh(
        cat_labels_short, vals, left=bottom,
        color=score_colors_bar[s], edgecolor="white",
        linewidth=0.6, label=f"Skor {s}",
    )
    for bar, v_val, bot in zip(bars_h, vals, bottom):
        if v_val > 3:
            ax_bar.text(
                bot + v_val / 2, bar.get_y() + bar.get_height() / 2,
                str(v_val), ha="center", va="center",
                fontsize=7.5, color="white", fontweight="bold",
            )
    bottom += vals

ax_bar.set_xlabel("Jumlah jawaban", fontsize=9)
ax_bar.set_title("Distribusi Skor per Kategori", fontsize=10.5, fontweight="bold")
ax_bar.legend(loc="lower right", fontsize=8, ncol=5, framealpha=0.9)
ax_bar.spines["top"].set_visible(False)
ax_bar.spines["right"].set_visible(False)

plt.tight_layout()
out3 = save(fig, "fig3_distribusi_skor.png")


# ─── 6. Gambar 4 – Heatmap per responden ──────────────────────────────────────

data_matrix = df_score.values.astype(float)

fig, ax = plt.subplots(figsize=(14, 10))
im = ax.imshow(data_matrix, cmap=plt.cm.RdYlGn, vmin=1, vmax=5, aspect="auto")

for i in range(n_resp):
    for j in range(len(LIKERT_COLS)):
        val        = int(data_matrix[i, j])
        text_color = "white" if val <= 2 or val == 5 else "black"
        ax.text(
            j, i, str(val),
            ha="center", va="center",
            fontsize=7.5, color=text_color, fontweight="bold",
        )

ax.set_xticks(range(len(LIKERT_COLS)))
ax.set_xticklabels(LIKERT_COLS, fontsize=9)
ax.set_yticks(range(n_resp))
ax.set_yticklabels([f"R{i+1:02d}" for i in range(n_resp)], fontsize=7.5)
ax.set_title(
    f"Heatmap Skor UAT per Responden\n(R01–R{n_resp:02d}, warna: merah=rendah, hijau=tinggi)",
    fontsize=11, fontweight="bold", pad=12,
)

for xpos in [3.5, 7.5, 11.5]:
    ax.axvline(xpos, color="white", linewidth=2)

cbar = fig.colorbar(im, ax=ax, fraction=0.015, pad=0.01)
cbar.set_label("Skor", fontsize=9)
cbar.set_ticks([1, 2, 3, 4, 5])

for k, v in KATEGORI.items():
    cols_idx = [LIKERT_COLS.index(c) for c in v["cols"]]
    mid = np.mean(cols_idx)
    ax.text(
        mid, -1.3,
        f"Kat. {k}\n{v['label'].replace(chr(10), ' ')}",
        ha="center", va="center",
        fontsize=8, fontweight="bold", color=v["color"],
    )

plt.tight_layout()
out4 = save(fig, "fig4_heatmap.png")


# ─── 7. Tabel statistik → Excel ───────────────────────────────────────────────

rows_stat = []
for col in LIKERT_COLS:
    s       = df_score[col]
    cat_key = col[0]
    rows_stat.append({
        "Kode"       : col,
        "Kategori"   : f"Kat. {cat_key} – {KATEGORI[cat_key]['label'].replace(chr(10), ' ')}",
        "Pernyataan" : PERTANYAAN[col].replace("\n", " "),
        "N"          : int(s.count()),
        "Min"        : int(s.min()),
        "Maks"       : int(s.max()),
        "Rata-rata"  : round(s.mean(), 2),
        "Std. Dev."  : round(s.std(), 2),
        "Skor 1 (%)" : round((s == 1).mean() * 100, 1),
        "Skor 2 (%)" : round((s == 2).mean() * 100, 1),
        "Skor 3 (%)" : round((s == 3).mean() * 100, 1),
        "Skor 4 (%)" : round((s == 4).mean() * 100, 1),
        "Skor 5 (%)" : round((s == 5).mean() * 100, 1),
    })

df_stat = pd.DataFrame(rows_stat)

for k, v in KATEGORI.items():
    sub = df_score[v["cols"]].values.flatten()
    df_stat = pd.concat([df_stat, pd.DataFrame([{
        "Kode"       : f"Rata-rata Kat. {k}",
        "Kategori"   : f"Kat. {k} – {v['label'].replace(chr(10), ' ')}",
        "Pernyataan" : "",
        "N": "", "Min": "", "Maks": "",
        "Rata-rata"  : round(sub.mean(), 2),
        "Std. Dev."  : "",
        "Skor 1 (%)": "", "Skor 2 (%)": "",
        "Skor 3 (%)": "", "Skor 4 (%)": "", "Skor 5 (%)": "",
    }])], ignore_index=True)

df_stat = pd.concat([df_stat, pd.DataFrame([{
    "Kode": "OVERALL", "Kategori": "Semua Kategori", "Pernyataan": "",
    "N": "", "Min": "", "Maks": "",
    "Rata-rata": overall, "Std. Dev.": "",
    "Skor 1 (%)": "", "Skor 2 (%)": "",
    "Skor 3 (%)": "", "Skor 4 (%)": "", "Skor 5 (%)": "",
}])], ignore_index=True)

out_xlsx = os.path.join(OUT_DIR, "tabel_statistik_uat.xlsx")
df_stat.to_excel(out_xlsx, index=False, sheet_name="Statistik UAT")

# ── Format Excel ──────────────────────────────────────────────────────────────
wb = load_workbook(out_xlsx)
ws = wb.active

header_fill = PatternFill("solid", fgColor="1F497D")
header_font = Font(bold=True, color="FFFFFF", size=10)

cat_fills = {
    "A": PatternFill("solid", fgColor="DDEEFF"),
    "B": PatternFill("solid", fgColor="DDFFEE"),
    "C": PatternFill("solid", fgColor="FFF3CC"),
    "D": PatternFill("solid", fgColor="FFE8D6"),
}
subtotal_fill = PatternFill("solid", fgColor="D9E1F2")
overall_fill  = PatternFill("solid", fgColor="E2EFDA")
bold_font     = Font(bold=True, size=10)
normal_font   = Font(size=10)
thin_border   = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

for cell in ws[1]:
    cell.fill      = header_fill
    cell.font      = header_font
    cell.alignment = center_align
    cell.border    = thin_border

col_widths = [16, 30, 52, 5, 5, 6, 11, 10, 10, 10, 10, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 28

for row in ws.iter_rows(min_row=2):
    kode        = str(row[0].value or "")
    cat_key     = kode[0] if kode and kode[0] in "ABCD" else None
    is_subtotal = kode.startswith("Rata-rata Kat.")
    is_overall  = kode == "OVERALL"

    for cell in row:
        cell.font   = bold_font if (is_subtotal or is_overall) else normal_font
        cell.border = thin_border

        if is_overall:
            cell.fill = overall_fill
        elif is_subtotal:
            cell.fill = subtotal_fill
        elif cat_key and cat_key in cat_fills:
            cell.fill = cat_fills[cat_key]

        cell.alignment = (
            left_align if cell.column == 3 else center_align
        )

ws.freeze_panes      = "A2"
ws.auto_filter.ref   = ws.dimensions

wb.save(out_xlsx)
print(f"  Disimpan: {out_xlsx}")

print()
print("✅ Selesai! Semua output tersimpan di:")
print(f"   {UAT_DIR}")
print()
print("  File yang dihasilkan:")
for f in [out1, out2, out3, out4, out_xlsx]:
    print(f"    {os.path.basename(f)}")